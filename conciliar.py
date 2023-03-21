from tkinter.filedialog import askopenfilename as askf
from openpyxl import Workbook
from subprocess import Popen
from tkinter import Tk
import pandas as pd
import numpy as np


class Archivo:

    def __init__ (self):
        # self.file1 = askf(title= 'Abre la conciliacion de bancos')
        # self.file2 = askf(title= 'Abre la conciliacion de libros')
        # self.file3 = askf(title= 'Abre la Tarjeta de credito Banco de bogota')
        # self.file4 = askf(title= 'Abre la Tarjeta de credito Davivienda')
        self.file1 = "D:\Team comunicaciones\conciliacion\\archivos prueba ypatia\Consolidado.xlsx"
        self.file2 = "D:\Team comunicaciones\conciliacion\\archivos prueba ypatia\libros.xlsx"
        self.file3 = "D:\Team comunicaciones\conciliacion\\archivos prueba ypatia\VENTAS CON TARJETA BOGOTA 9314.xlsx.csv"
        self.file4 = "D:\Team comunicaciones\conciliacion\\archivos prueba ypatia\VENTAS CON TARJETA DAVIVIENDA.xlsx"


    def data(self):
        self.data1 = pd.read_excel(self.file1, header= None)
        self.data1 = np.asarray(self.data1)
        self.data1 = self.data1.tolist()
        self.data2 = pd.read_excel(self.file2, header= None)
        self.data2 = np.asarray(self.data2)
        self.data2 = self.data2.tolist()
        self.tc_bogota = pd.read_excel(self.file3, header= None)
        self.tc_bogota = np.asarray(self.tc_bogota)
        self.tc_bogota = self.tc_bogota.tolist()
        self.tc_davivienda = pd.read_excel(self.file4, header= None)
        self.tc_davivienda = np.asarray(self.tc_davivienda)
        self.tc_davivienda = self.tc_davivienda.tolist()
    
    def identify (self):
        temp1 = self.data1
        temp2 = self.data2
        if 'Banco' in temp1[0][0]: 
            self.data2 = temp1
            self.data1 = None
        else: 
            self.data1 = temp1
            self.data2 = None
        if 'Banco' in temp2[0][0]: 
            self.data2 = temp2
        else: 
            self.data1 = temp2
    
    def conciliar(self):
        temp1 = self.data1
        temp2 = self.data2
        self.libros_true = []
        self.libros_false = []
        self.bancos_true = []
        self.bancos_false = []
        for i in temp1:
            if str(i[9]).replace(" ","") != "nan" and str(i[9]).replace(" ","") != 'autorizacion': 
                self.libros_false.append(i)
                continue
            date1= i[3]
            value1= i[8]
            bank1= i[1]
            check = True
            for j in range(0, len(temp2)):
                date2= temp2[j][1]
                value2= temp2[j][3]
                bank2= temp2[j][0]
                check_date = date1 == date2
                check_value = value1 == value2
                check_bank = bank1 == bank2
                if check_date and check_value and check_bank:
                    self.libros_true.append(i)
                    self.bancos_true.append(temp2[j])
                    temp2.pop(j)
                    check = False
                    break
            if check: self.libros_false.append(i)
        self.bancos_false = temp2

    def filters(self):
        temp = []
        for i in self.bancos_false:
            if str(i[4]).replace(" ","") == "nan" or str(i[4]).replace(" ","") == "Notas":
                temp.append(i)
            else: self.bancos_true.append(i)
        self.bancos_false = temp

    def bank_bogota_code(self):
        self.bancos_false[0].append('Codigo')
        for i in range(1,len(self.bancos_false)):
            if 'Deposito electronico ventas con tarjetas' in self.bancos_false[i][2]:
                code = str(self.bancos_false[i][2]).replace("Deposito electronico ventas con tarjetas","")
                code = code.replace(" ","")
                self.bancos_false[i].append(code)
            else:
                self.bancos_false[i].append('')
    
    def bank_davivienda_code(self):
        for i in range(1,len(self.bancos_false)):
            descripcion = str(self.bancos_false[i][2]).lower().replace(" ","")
            if 'ventasnetas' in descripcion:
                code = "davivienda"
                self.bancos_false[i][5] = code

    
    def move_data(self, position, origen, receive, add = None):
        data= origen[position]
        if add != None:
            data = data + add
        receive.append(data)
        origen.pop(position)
    
    def check_data(self, data1, data2):
        check = False
        if len(data1) == len(data2):
            check = True
            for i in range (0,len(data1)):
                if data1[i]  != data2[i]: check = False
        return check             
    
    def tc_bogota_procces(self):
        self.tc_bogota_true = []
        self.tc_bogota_false = self.tc_bogota
        self.libro_tarjetas_false = []
        tempdata= []

        for i_l_f in range(0, len(self.libros_false)):
            check_data_libro= True
            data_libro = self.libros_false[i_l_f]
            if data_libro[0] == 'Cuenta': continue
            autorizacion = str(data_libro[9])
            valor_libro = int(data_libro[8])
            if str(autorizacion).replace(" ","") != 'nan' and autorizacion!= 'autorizacion':
                for i_tb_f in range (0, len(self.tc_bogota_false)):
                    data_tc_bogota = self.tc_bogota_false[i_tb_f]
                    autorizacion_tc = str(data_tc_bogota[9])
                    franquicia = data_tc_bogota[13]
                    codigo_tc = data_tc_bogota[7]
                    total = data_tc_bogota[18]
                    comision = data_tc_bogota[20]
                    reterenta = data_tc_bogota[22]
                    reteiva = data_tc_bogota[23]
                    reteica = data_tc_bogota[24]
                    neto = data_tc_bogota[25]
                    metadata= [total, comision, reterenta, reteiva, reteica, neto]

                    if autorizacion == autorizacion_tc:
                        check_data_libro=False
                        if valor_libro == neto:
                            self.move_data(i_tb_f, self.tc_bogota_false, self.tc_bogota_true)
                            tempdata.append(data_libro+[codigo_tc,franquicia])
                            break
                        else:
                            self.libro_tarjetas_false.append(data_libro+metadata)
                            break
            if check_data_libro: tempdata.append(data_libro)        
        self.libros_false = tempdata

    def tc_bogota_procces2(self):
        tempdata = []
        for i_b_f in range(0, len(self.bancos_false)):
            check_data_banco = True
            data_banco = self.bancos_false[i_b_f]
            codigo_banco = data_banco[5]
            value_banco = data_banco[3]
            total_values_libros = 0
            posiciones_libros = []
            for i_l_f in range(0, len(self.libros_false)):
                data_libro = self.libros_false[i_l_f]
                if len(data_libro) <11: continue
                codigo_libro = str(data_libro[10])
                value_libro = data_libro[8]
                if codigo_banco == codigo_libro:
                    if value_banco == value_libro:
                        self.move_data(i_l_f, self.libros_false, self.libros_true)
                        self.bancos_true.append(data_banco)
                        check_data_banco = False
                        break
                    total_values_libros += value_libro
                    posiciones_libros.append(i_l_f)
                    if value_banco == total_values_libros:
                        posiciones_libros.sort(reverse=True)
                        for posicion in posiciones_libros:
                            self.move_data(posicion, self.libros_false, self.libros_true)
                        self.bancos_true.append(data_banco)
                        check_data_banco = False
                        break
            if check_data_banco:
                tempdata.append(data_banco)
        self.bancos_false = tempdata
    
    def tc_davivienda_procces(self):
        self.tc_davivienda_true = []
        self.tc_davivienda_false = self.tc_davivienda
        tempdata= []
        self.libro_tc_davivienda_false =[]

        for i_l_f in range(0, len(self.libros_false)):
            check_data_libro= True
            data_libro = self.libros_false[i_l_f]
            if data_libro[0] == 'Cuenta': continue
            autorizacion = str(data_libro[9])
            valor_libro = int(data_libro[8])
            if str(autorizacion).replace(" ","") != 'nan' and autorizacion!= 'autorizacion':
                for i_td_f in range (0, len(self.tc_davivienda_false)):
                    data_tc_davivienda = self.tc_davivienda_false[i_td_f]
                    autorizacion_tc = str(data_tc_davivienda[8])
                    terminal = str(data_tc_davivienda[7])
                    ruta = str(data_tc_davivienda[3])

                    total = data_tc_davivienda[9]
                    iva = data_tc_davivienda[10]
                    ipoconsumo = data_tc_davivienda[11]
                    comision = data_tc_davivienda[13]
                    retefuente = data_tc_davivienda[14]
                    reteiva = data_tc_davivienda[15]
                    reteica = data_tc_davivienda[16]
                    neto = data_tc_davivienda[17]
                    neto= str(neto).replace("$ ","").replace(",00","").replace(".","")
                    metadata= [total, iva, ipoconsumo, comision, retefuente, reteiva, reteica, neto]

                    if autorizacion == autorizacion_tc:
                        check_data_libro=False
                        if valor_libro == int(neto):
                            self.move_data(i_td_f, self.tc_davivienda_false, self.tc_davivienda_true)
                            tempdata.append(data_libro+[terminal,ruta])
                            break
                        else:
                            self.libro_tc_davivienda_false.append(data_libro+metadata)
                            break
            if check_data_libro: tempdata.append(data_libro)        
        self.libros_false = tempdata
                    


    def tc_davivienda_procces2(self):
        tempdata = []
        for i_b_f in range(0, len(self.bancos_false)):
            check_data_banco = True
            data_banco = self.bancos_false[i_b_f]
            banco = data_banco[0]
            codigo_banco = data_banco[5]
            value_banco = data_banco[3]
            total_values_libros = 0
            posiciones_libros = []
            if banco != '3996699969-73' or codigo_banco != 'davivienda': 
                tempdata.append(data_banco)
                continue
            for i_l_f in range(0, len(self.libros_false)):
                data_libro = self.libros_false[i_l_f]
                if len(data_libro) <11: continue
                # codigo_libro = str(data_libro[10])
                value_libro = data_libro[8]
                # if codigo_banco == codigo_libro:
                if value_banco == value_libro:
                    self.move_data(i_l_f, self.libros_false, self.libros_true)
                    self.bancos_true.append(data_banco)
                    check_data_banco = False
                    break
        #         total_values_libros += value_libro
        #         posiciones_libros.append(i_l_f)
        #         if value_banco == total_values_libros:
        #             posiciones_libros.sort(reverse=True)
        #             for posicion in posiciones_libros:
        #                 self.move_data(posicion, self.libros_false, self.libros_true)
        #             self.bancos_true.append(data_banco)
        #             check_data_banco = False
        #             break
            if check_data_banco:
                tempdata.append(data_banco)
        self.bancos_false = tempdata
    
    def coincide_fecha(self):
        temp1 = self.libros_false
        temp2 = self.bancos_false
        tempdata =[]
        self.libros_fecha = []
        self.bancos_fecha = []
        for i in temp1:
            if str(i[9]).replace(" ","") != "nan" and str(i[9]).replace(" ","") != 'autorizacion': 
                tempdata.append(i)
                continue
            date1= i[3]
            value1= i[8]
            bank1= i[1]
            check = True
            for j in range(0, len(temp2)):
                date2= temp2[j][1]
                value2= temp2[j][3]
                bank2= temp2[j][0]
                check_date = date1 == date2
                check_value = value1 == value2
                check_bank = bank1 == bank2
                if check_date and check_value:
                    self.libros_fecha.append(i)
                    self.bancos_fecha.append(temp2[j])
                    temp2.pop(j)
                    check = False
                    break
            if check: tempdata.append(i)
        self.bancos_false = temp2
        self.libros_false = tempdata

    def coincide_banco(self):
        temp1 = self.libros_false
        temp2 = self.bancos_false
        tempdata =[]
        self.libros_banco = []
        self.bancos_banco = []
        for i in temp1:
            if str(i[9]).replace(" ","") != "nan" and str(i[9]).replace(" ","") != 'autorizacion': 
                tempdata.append(i)
                continue
            date1= i[3]
            value1= i[8]
            bank1= i[1]
            check = True
            for j in range(0, len(temp2)):
                date2= temp2[j][1]
                value2= temp2[j][3]
                bank2= temp2[j][0]
                check_date = date1 == date2
                check_value = value1 == value2
                check_bank = bank1 == bank2
                if check_bank and check_value:
                    self.libros_banco.append(i)
                    self.bancos_banco.append(temp2[j])
                    temp2.pop(j)
                    check = False
                    break
            if check: tempdata.append(i)
        self.bancos_false = temp2
        self.libros_false = tempdata





  
    def print_excel(self):
        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = 'Si_libro'
        for fila in self.libros_true: sheet1.append(fila)
        wb.create_sheet(index=1, title= 'Si_banco')
        sheet2 = wb.get_sheet_by_name('Si_banco')
        for fila in self.bancos_true: sheet2.append(fila)
        wb.create_sheet(index=2, title= 'no_libro')
        sheet3 = wb.get_sheet_by_name('no_libro')
        for fila in self.libros_false: sheet3.append(fila)
        wb.create_sheet(index=3, title= 'no_banco')
        sheet4 = wb.get_sheet_by_name('no_banco')
        for fila in self.bancos_false: sheet4.append(fila)
        wb.create_sheet(index=4, title= 'no_tc_bogota')
        sheet5 = wb.get_sheet_by_name('no_tc_bogota')
        for fila in self.tc_bogota_false: sheet5.append(fila)
        wb.create_sheet(index=5, title= 'si_tc_bogota')
        sheet6 = wb.get_sheet_by_name('si_tc_bogota')
        for fila in self.tc_bogota_true: sheet6.append(fila)
        wb.create_sheet(index=6, title= 'corregir libro')
        sheet7 = wb.get_sheet_by_name('corregir libro')
        for fila in self.libro_tarjetas_false: sheet7.append(fila)
        wb.create_sheet(index=7, title= 'si_tc_davivienda')
        sheet8 = wb.get_sheet_by_name('si_tc_davivienda')
        for fila in self.tc_davivienda_true: sheet8.append(fila)
        wb.create_sheet(index=8, title= 'no_tc_davivienda')
        sheet9 = wb.get_sheet_by_name('no_tc_davivienda')
        for fila in self.tc_davivienda_false: sheet9.append(fila)
        wb.create_sheet(index=9, title= 'corregir libroD')
        sheet10 = wb.get_sheet_by_name('corregir libroD')
        for fila in self.libro_tc_davivienda_false: sheet10.append(fila)
        wb.create_sheet(index=10, title= 'posible libro fecha')
        sheet11 = wb.get_sheet_by_name('posible libro fecha')
        for fila in self.libros_fecha: sheet11.append(fila)
        wb.create_sheet(index=11, title= 'posible banco fecha')
        sheet12 = wb.get_sheet_by_name('posible banco fecha')
        for fila in self.bancos_fecha: sheet12.append(fila)
        wb.create_sheet(index=12, title= 'posible libro cuenta')
        sheet13 = wb.get_sheet_by_name('posible libro cuenta')
        for fila in self.libros_banco: sheet13.append(fila)
        wb.create_sheet(index=13, title= 'posible banco cuenta')
        sheet14 = wb.get_sheet_by_name('posible banco cuenta')
        for fila in self.bancos_banco: sheet14.append(fila)
        wb.save('Consolidado.xlsx')
        p = Popen("openExcel.bat")
        stdout, stderr = p.communicate()


           
def run():
    conciliar = Archivo()
    conciliar.data()
    conciliar.identify()
    conciliar.conciliar()
    conciliar.filters()
    conciliar.bank_bogota_code()
    conciliar.bank_davivienda_code()
    conciliar.tc_bogota_procces()
    conciliar.tc_davivienda_procces()
    conciliar.tc_bogota_procces2()
    conciliar.tc_davivienda_procces2()
    conciliar.coincide_fecha()
    conciliar.coincide_banco()
    conciliar.print_excel()

#run()