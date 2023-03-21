from tkinter import Tk
from tkinter.filedialog import askopenfilename as askf
import pandas as pd
import numpy as np
from openpyxl import Workbook
from subprocess import Popen
import datetime

mes="09"
año="2022"

consolidado=[]
bancos=[
    ['11100505', 'BANCOLOMBIA', '325-450229-55'],
    ['11100510', 'BANCO DE BOGOTA', '538-0393-14'],
    ['11100515', 'BANCO DAVIVIENDA CORRIENTE', '3996699969-73'],
    ['11100520', 'COLPATRIA MEDELLIN TAT', '6921005278'],
    ['11200505', 'BANCOLOMBIA', '325-455868-34'],
    ['11200524', 'BOGOTA (CALI POPAYAN)', '538-1765-87'],
    ['11200530', 'CONVENIO URABA 11418', '645-626385-88'],
    ['11200545', 'CONVENIO CHOCO 11419', '536-629632-21'],
    ['11200558', 'CONVENIO CAUCASIA 11423', '371-000005-45'],
    ['11200575', 'RECAUDOS TAT', '551-798737-44'],
    ['11200586', 'CONVENIO EJE CAFETERO 11422', '859-000119-92'],
    ['11200593', 'CONVENIO BOYACA 11421', '358-972792-13'],
    ['11200595', 'BANCO AGRARIO', '4-1331-301395-5'],
    ['11200598', 'CONVENIO OR OCC 11420', '024-470946-09'],
    ['11100540', '7187 DAVIVIENDA', '399669997187'],
    ['11200521', 'CAJA SOCIAL AHORROS', '24115476024'],
    ['11100545', 'CAJA SOCIAL CORRIENTE', '21004128844']
]

def pickfile(banco):
    print("seleccione archivo para "+ banco[1]+ " cuenta: "+ banco[0]+ "  "+ banco[2])
    filename = askf()
    if filename == '': filename = None 
    banco.append(filename)

def bancolombia(banco):
    data=leerExcel(banco)
    for i in data:
        x= str(i[6]).replace(" ","")
        if (x != "SALDODIA" and x != "SALDOINICIAL" and x != "SALDOFINAL"):
            cuenta= i[1]
            fecha= str(i[0])[6:8]+"/"+str(i[0])[4:6]+"/"+str(i[0])[0:4]
            descripcion = i[6]
            debito= i[8]
            concatenado = str(debito)+"-"+str(fecha)+"-"+str(cuenta)
            tempData= cuenta, fecha, descripcion, debito, concatenado
            consolidado.append(tempData)

def colpatria(banco):
    data = leerExcel(banco)
    cuenta = banco[2]
    start = False
    for i in data:
        if start:
            fecha = i[6]
            descripcion = i[8]
            deposito= i[9]
            if str(deposito).replace(" ","") == "nan": deposito=0
            retiro= i[10]
            if str(retiro).replace(" ","") == "nan": retiro=0
            if int(deposito) > int(retiro): debito = deposito
            else: debito = -retiro
            concatenado = str(debito)+"-"+str(fecha)+"-"+str(cuenta)
            tempData= cuenta, fecha, descripcion, debito, concatenado
            consolidado.append(tempData)
        if str(i[0]).replace(" ","") == "Naturaleza": start = True

def cajaSocial(banco):
    data = leerExcel(banco)
    cuenta = banco[2]
    start = False
    for i in data:
        if start:
            fecha = i[4]
            descripcion = i[2]
            debito = i[0] 
            concatenado = str(debito)+"-"+str(fecha)+"-"+str(cuenta)
            tempData= cuenta, fecha, descripcion, debito, concatenado
            consolidado.append(tempData)
        if str(i[0]).replace(" ","") == "VALOR_TRANSACCION": start = True

def agrario(banco):
    data = leerExcel(banco)
    cuenta= banco[2]
    start= False
    for i in data:
        if start:
            fecha= i[0]
            descripcion= i[2]
            deposito= i[3]
            if str(deposito).replace(" ","") == "nan": deposito=0
            retiro= i[4]
            if str(retiro).replace(" ","") == "nan": retiro=0
            if int(deposito) > int(retiro): debito = deposito
            else: debito = -retiro
            concatenado = str(debito)+"-"+str(fecha)+"-"+str(cuenta)
            tempData= cuenta, fecha, descripcion, debito, concatenado
            consolidado.append(tempData)
        if str(i[0]).replace(" ","") == "Fecha": start = True

def davivienda(banco):
    data = leerExcel(banco)
    cuenta= banco[2]
    start= False
    for i in data:
        if start:
            fecha = i[0]
            descripcion =i[2]
            debito= i[7]
            concatenado = str(debito)+"-"+str(fecha)+"-"+str(cuenta)
            tempData= cuenta, fecha, descripcion, debito, concatenado
            consolidado.append(tempData)
        if str(i[0]).replace(" ","") == "FechadeSistema": start=True

def bancoBogota(banco):
    data=leerExcel(banco)
    cuenta= banco[2]
    start= False
    for i in data:
        if start:
            fecha= i[0]
            if (str(type(fecha)) != "<class 'datetime.datetime'>"):
                if (int(fecha) < 32 and int(fecha)> 0): fecha = año+"/"+mes+"/"+str(fecha)
            descripcion= i[1]
            deposito= i[3]
            if str(deposito).replace(" ","") == "nan": deposito=0
            retiro= i[4]
            if str(retiro).replace(" ","") == "nan": retiro=0
            if int(deposito) > int(retiro): debito = deposito
            else: debito = -retiro
            concatenado = str(debito)+"-"+str(fecha)+"-"+str(cuenta)
            tempData= cuenta, fecha, descripcion, debito, concatenado
            consolidado.append(tempData)
        if str(i[0]).replace(" ","") == "Fecha": start=True

def leerExcel(banco):
    data=[]
    if banco[3] != None:
        fileExcel= pd.read_excel(banco[3], header=None)
        data = np.asarray(fileExcel)
    return data

def generarExcel():
    wb = Workbook()
    hoja = wb.active
    hoja.append(('Cuenta Banco', 'Fecha', 'Descripcion', 'Debitos'))
    celda=2
    for fila in consolidado:
        if (str(fila[2]).replace(" ","") != "nan"):
            hoja.append(fila)
            hoja.cell(row=celda, column=2).number_format="0"
            celda+=1


    wb.save('Consolidado.xlsx')

    p = Popen("openExcel.bat")
    stdout, stderr = p.communicate()

#try:
if True:
    for banco in bancos:
        pickfile(banco)

    for banco in bancos:
        if banco[3] != None:
            if banco[0] == '11100505' : bancolombia(banco)
            if banco[0] == '11200505' : bancolombia(banco)
            if banco[0] == '11200530' : bancolombia(banco)
            if banco[0] == '11200545' : bancolombia(banco)
            if banco[0] == '11200558' : bancolombia(banco)
            if banco[0] == '11200575' : bancolombia(banco)
            if banco[0] == '11200586' : bancolombia(banco)
            if banco[0] == '11200593' : bancolombia(banco)
            if banco[0] == '11200598' : bancolombia(banco)
            if banco[0] == '11100510' : bancoBogota(banco)
            # # if banco[0] == '11200524' : sin movimiento()
            if banco[0] == '11100515' : davivienda(banco)
            if banco[0] == '11100540' : davivienda(banco)
            if banco[0] == '11200595' : agrario(banco)
            if banco[0] == '11200521' : cajaSocial(banco)
            if banco[0] == '11100545' : cajaSocial(banco)
            if banco[0] == '11100520' : colpatria(banco)



    generarExcel()

# except Exception as e:
#     print(e)
#     while True:
#         pass



# fileExcel = pd.read_excel(filename)
# numbers =np.asarray(fileExcel)
# filename2 = askf()






# input("Numero de banco deseado: ")





# bancos = pd.read_excel('bancos.xlsx')
# wb = Workbook()
# ws= wb.active
# ws['A1'] = 4